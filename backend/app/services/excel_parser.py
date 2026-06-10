"""
Excel 周报解析服务
将运营周报 Excel 文件解析为数据库模型数据（营业额/套餐/运营指标/分析意见）

Excel 格式约定：
- Sheet "营业额情况": 按门店分行，列 A=门店名, B=指标类型, C/D=两周数据, E=环比
  第1行: 标题行 (门店/日期, 空, 5月第三周, 5月第四周, 环比)
  第2行: 日期行 (空, 空, 5.18-24号, 5.25-31号, 空)
  第3行起: 每7行一个门店 (店名/营业额/美团/抖音/到店人数/接待桌数/桌均人数/人均)
- Sheet "套餐数据": 左右两组(上周/本周)，每门店多行商品，末行汇总
  第1行: 日期 (上周日期, 空, ..., 本周日期, ...)
  第2行: 表头 (核销门店, 商品名称, 美团购买, 美团核销券数, 抖音核销券数, ...)
  第3行起: 数据行，门店行在A列，商品行在A列为空
- Sheet "XX店": 单门店运营指标表
  第1行: 门店名称
  第2行: 分析意见 (B列)
  第3行: 表头 (项目, 上周日期, 本周日期, 环比)
  第4行起: 指标行 (美团人气榜/点评热门榜/美团星级/曝光次数/访问次数/...)
"""

from __future__ import annotations

import math
import re
from dataclasses import dataclass, field
from datetime import date
from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.services.store_name_utils import find_best_matching_store


# ==================== 数据容器 ====================

@dataclass
class ParsedData:
    """解析结果容器"""
    revenues: list[dict] = field(default_factory=list)
    packages: list[dict] = field(default_factory=list)
    metrics: list[dict] = field(default_factory=list)
    analyses: list[dict] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ==================== 辅助函数 ====================

def _safe_float(val) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, float) and math.isnan(val):
        return None
    if isinstance(val, str):
        val = val.strip()
        if val in ("/", "", "持平", "未上榜", "NaN", "nan", "下滑一名"):
            return None
        m = re.match(r'^(-?\d+\.?\d*)%$', val)
        if m:
            return float(m.group(1)) / 100
        try:
            return float(val)
        except ValueError:
            return None
    return float(val) if isinstance(val, (int, float)) else None


def _safe_int(val) -> Optional[int]:
    f = _safe_float(val)
    return int(f) if f is not None else None


def _parse_short_date(s: str) -> Optional[date]:
    """解析 '5.24' 或 '5.31' 格式为 date(2026, 5, 24)"""
    if not s:
        return None
    s = s.strip().rstrip('号')
    m = re.match(r'^(\d{1,2})\.(\d{1,2})$', s)
    if not m:
        return None
    try:
        return date(2026, int(m.group(1)), int(m.group(2)))
    except ValueError:
        return None


def _parse_date_range_str(s: str) -> tuple[Optional[date], Optional[date]]:
    """解析 '5.18-24号' 或 '5.24-31号' 或 '5.24-31' 为 (start, end)"""
    if not s:
        return None, None
    s = s.strip().rstrip('号')
    m = re.match(r'^(\d{1,2})\.(\d{1,2})-(\d{1,2})$', s)
    if m:
        month = int(m.group(1))
        start = _parse_short_date(f'{month}.{m.group(2)}')
        end = _parse_short_date(f'{month}.{m.group(3)}')
        return start, end
    return None, None


def _get_cell_str(ws: Worksheet, row: int, col: int) -> str:
    val = ws.cell(row, col).value
    return str(val).strip() if val is not None else ""


# ==================== 营业额解析 ====================

def _parse_revenue_sheet(ws: Worksheet, store_name_map: dict[str, UUID]) -> tuple[list[dict], list[str], tuple[Optional[date], Optional[date]]]:
    revenues = []
    errors = []

    # 第2行: C列=上周日期范围, D列=本周日期范围
    week1_str = _get_cell_str(ws, 2, 3)
    week2_str = _get_cell_str(ws, 2, 4)
    w1_start, w1_end = _parse_date_range_str(week1_str)
    w2_start, w2_end = _parse_date_range_str(week2_str)
    period_start = w1_start
    period_end = w2_end

    if not w1_start or not w1_end:
        errors.append(f"营业额表: 无法解析上周日期范围 '{week1_str}'")
    if not w2_start or not w2_end:
        errors.append(f"营业额表: 无法解析本周日期范围 '{week2_str}'")

    def _read_revenue_row(start_row: int, col: int) -> dict:
        """从指定起始行和列读取一个门店的营业额数据"""
        total = _safe_float(ws.cell(start_row, col).value) or 0
        mt = _safe_float(ws.cell(start_row + 1, col).value) or 0
        dy = _safe_float(ws.cell(start_row + 2, col).value) or 0
        visitors = _safe_int(ws.cell(start_row + 3, col).value) or 0
        tables = _safe_int(ws.cell(start_row + 4, col).value) or 0
        avg_people = _safe_float(ws.cell(start_row + 5, col).value) or 0
        avg_capita = _safe_float(ws.cell(start_row + 6, col).value) or 0
        other = max(0, total - mt - dy)
        return {
            "total_revenue": round(total, 2),
            "meituan_revenue": round(mt, 2),
            "douyin_revenue": round(dy, 2),
            "other_revenue": round(other, 2),
            "visitor_count": visitors,
            "table_count": tables,
            "avg_people_per_table": round(avg_people, 4),
            "avg_per_capita": round(avg_capita, 2),
        }

    # 第3行起，每7行一个门店
    row = 3
    while row <= ws.max_row:
        name = ws.cell(row, 1).value
        if not name or not isinstance(name, str) or not name.strip():
            row += 1
            continue
        name = name.strip()
        sid, sim, matched_name = find_best_matching_store(name, store_name_map, threshold=0.6)
        if not sid:
            errors.append(f"营业额表: 门店 '{name}' 未匹配到系统门店（最佳相似度={sim:.2f}）")
            row += 7
            continue
        if sim < 1.0:
            errors.append(f"营业额表: 门店 '{name}' 模糊匹配到 '{matched_name}'（相似度={sim:.2f}）")

        # 读取上周数据 (C列)
        if w1_end:
            week1_data = _read_revenue_row(row, 3)
            revenues.append({
                "store_id": sid,
                "record_date": w1_end,
                **week1_data,
            })

        # 读取本周数据 (D列)
        if w2_end:
            week2_data = _read_revenue_row(row, 4)
            revenues.append({
                "store_id": sid,
                "record_date": w2_end,
                **week2_data,
            })

        row += 7

    return revenues, errors, (period_start, period_end)


# ==================== 套餐解析 ====================

def _parse_package_sheet(ws: Worksheet, store_name_map: dict[str, UUID]) -> tuple[list[dict], list[str]]:
    packages = []
    errors = []

    # 第1行: A列=上周日期, F列=本周日期
    prev_str = _get_cell_str(ws, 1, 1)
    curr_str = _get_cell_str(ws, 1, 6)
    prev_start, prev_end = _parse_date_range_str(prev_str)
    curr_start, curr_end = _parse_date_range_str(curr_str)

    if not prev_start or not prev_end:
        errors.append(f"套餐表: 无法解析上周日期范围 '{prev_str}'")
    if not curr_start or not curr_end:
        errors.append(f"套餐表: 无法解析本周日期范围 '{curr_str}'")

    def _read_package_row(row: int, product_col: int, buy_col: int, verify_col: int, dy_col: int) -> Optional[dict]:
        """从指定列读取一行套餐数据"""
        product = ws.cell(row, product_col).value
        if not product or not isinstance(product, str) or not product.strip() or product.strip().lower() in ("汇总", "0", "nan"):
            return None
        return {
            "product_name": product.strip(),
            "meituan_buy": _safe_int(ws.cell(row, buy_col).value) or 0,
            "meituan_verify": _safe_int(ws.cell(row, verify_col).value) or 0,
            "douyin_buy": 0,
            "douyin_verify": _safe_int(ws.cell(row, dy_col).value) or 0,
        }

    current_store_id = None
    row = 3  # 跳过表头行(第2行是"核销门店"表头)
    while row <= ws.max_row:
        # 检查A列是否是门店名
        a_val = ws.cell(row, 1).value
        if a_val and isinstance(a_val, str) and a_val.strip() and a_val.strip() not in ("汇总", "0"):
            name = a_val.strip()
            sid_p, sim_p, matched_p = find_best_matching_store(name, store_name_map, threshold=0.6)
            current_store_id = sid_p
            if not current_store_id:
                errors.append(f"套餐表: 门店 '{name}' 未匹配到系统门店（最佳相似度={sim_p:.2f}）")
            elif sim_p < 1.0:
                errors.append(f"套餐表: 门店 '{name}' 模糊匹配到 '{matched_p}'（相似度={sim_p:.2f}）")
            row += 1
            continue

        if current_store_id is None:
            row += 1
            continue

        # 读取上周数据: B=商品名, C=美团购买, D=美团核销, E=抖音核销
        if prev_start and prev_end:
            prev_pkg = _read_package_row(row, 2, 3, 4, 5)
            if prev_pkg:
                packages.append({
                    "store_id": current_store_id,
                    "period_start": prev_start,
                    "period_end": prev_end,
                    **prev_pkg,
                })

        # 读取本周数据: F=商品名, G=美团购买, H=美团核销, I=抖音核销
        if curr_start and curr_end:
            curr_pkg = _read_package_row(row, 6, 7, 8, 9)
            if curr_pkg:
                packages.append({
                    "store_id": current_store_id,
                    "period_start": curr_start,
                    "period_end": curr_end,
                    **curr_pkg,
                })

        row += 1

    return packages, errors


# ==================== 门店指标和分析解析 ====================

def _find_str(rows: list[tuple[str, object]], keywords: list[str]) -> Optional[str]:
    for name, val, _ in rows:
        if any(kw in name for kw in keywords):
            v = str(val).strip() if val is not None else ""
            return v if v and v not in ("/", "NaN", "nan") else None
    return None


def _find_float(rows: list[tuple[str, object]], keywords: list[str]) -> Optional[float]:
    for name, val, _ in rows:
        if any(kw in name for kw in keywords):
            return _safe_float(val)
    return None


def _find_int(rows: list[tuple[str, object]], keywords: list[str]) -> Optional[int]:
    return _safe_int(_find_float(rows, keywords))


def _extract_rank(s: Optional[str]) -> Optional[str]:
    if not s:
        return None
    m = re.search(r'第\d+名', s)
    return m.group(0) if m else (s if s not in ("未上榜", "/", "") else None)


def _extract_bad_keywords(text: Optional[str]) -> Optional[list[str]]:
    if not text:
        return None
    m = re.search(r'差评关键词[：:]([^\n]+)', text)
    if m:
        kws = [k.strip() for k in m.group(1).split('、') if k.strip()]
        return kws if kws else None
    return None


def _extract_goals(text: Optional[str]) -> list[str]:
    if not text:
        return []
    m = re.search(r'下周目标[^：:]*[：:](.*?)(?=\n|$)', text, re.DOTALL)
    if m:
        items = re.split(r'[，,；;\n]', m.group(1).strip())
        return [i.strip() for i in items if i.strip() and len(i.strip()) > 2]
    return []


def _parse_store_sheet(
    ws: Worksheet,
    store_name_map: dict[str, UUID],
    sheet_name: str,
    default_period: tuple[Optional[date], Optional[date]],
) -> tuple[list[dict], list[str], Optional[dict]]:
    """解析单个门店 sheet（同时读取上周B列和本周C列数据）"""
    metrics = []
    errors = []
    analysis = None

    title = ws.cell(1, 1).value or ""
    query_candidates = [str(title).strip(), sheet_name.strip()]
    sid = None
    for qname in query_candidates:
        sid, sim_s, matched_s = find_best_matching_store(qname, store_name_map, threshold=0.6)
        if sid:
            if sim_s < 1.0:
                errors.append(f"{sheet_name}: 模糊匹配到 '{matched_s}'（相似度={sim_s:.2f}）")
            break
    if not sid:
        errors.append(f"{sheet_name}: 未匹配到系统门店")
        return metrics, errors, analysis

    # 分析意见 (第2行 B列)
    opinion = ws.cell(2, 2).value
    if opinion and isinstance(opinion, str):
        opinion = opinion.strip()

    # 日期 (第3行): B列=上周日期范围, C列=本周日期范围
    date_b = _get_cell_str(ws, 3, 2)
    date_c = _get_cell_str(ws, 3, 3)
    prev_start, prev_end = _parse_date_range_str(date_b)
    curr_start, curr_end = _parse_date_range_str(date_c)

    if not prev_start or not prev_end:
        errors.append(f"{sheet_name}: 无法解析上周日期范围 '{date_b}'")
    if not curr_start or not curr_end:
        errors.append(f"{sheet_name}: 无法解析本周日期范围 '{date_c}'")

    # 使用解析到的日期，如果解析失败则回退到默认值
    if not prev_end:
        prev_end = default_period[0] or curr_end
    if not curr_end:
        curr_end = default_period[1] or date.today()

    # 收集指标行 (第4行起)
    # 门店 sheet 结构: A=指标名, B=上周, C=本周, D=环比
    rows_prev = []
    rows_curr = []
    for r in range(4, ws.max_row + 1):
        proj = ws.cell(r, 1).value
        if proj and isinstance(proj, str) and proj.strip():
            name = proj.strip()
            rows_prev.append((name, ws.cell(r, 2).value, r))
            rows_curr.append((name, ws.cell(r, 3).value, r))

    def _make_platform_metrics(rows_list, m_date, platform):
        """根据rows数据和日期创建指定平台的指标"""
        if platform == "meituan":
            return {
                "store_id": sid,
                "metric_date": m_date,
                "platform": "meituan",
                "ranking_name": _find_str(rows_list, ["美团人气榜"]),
                "ranking_position": _extract_rank(_find_str(rows_list, ["美团人气榜"])),
                "star_rating": _find_float(rows_list, ["美团星级"]),
                "impressions": _find_int(rows_list, ["曝光次数"]),
                "visits": _find_int(rows_list, ["访问次数"]),
                "purchases": _find_int(rows_list, ["购买人数"]),
                "new_favorites": _find_int(rows_list, ["新增收藏"]),
                "checkins": _find_int(rows_list, ["打卡人数"]),
                "scan_count": _find_int(rows_list, ["扫码人数"]),
                "product_impressions": _find_int(rows_list, ["商品曝光人数"]),
                "product_visits": _find_int(rows_list, ["商品访问人数"]),
                "product_purchases": _find_int(rows_list, ["商品购买人数"]),
                "new_reviews": _find_int(rows_list, ["新评价数"]),
                "new_bad_reviews": _find_int(rows_list, ["新中差评"]),
            }
        elif platform == "dianping":
            return {
                "store_id": sid,
                "metric_date": m_date,
                "platform": "dianping",
                "ranking_name": _find_str(rows_list, ["点评热门榜"]),
                "ranking_position": _extract_rank(_find_str(rows_list, ["点评热门榜"])),
                "star_rating": _find_float(rows_list, ["点评星级"]),
            }
        elif platform == "douyin":
            return {
                "store_id": sid,
                "metric_date": m_date,
                "platform": "douyin",
                "ranking_name": _find_str(rows_list, ["抖音人气榜"]),
                "ranking_position": _extract_rank(_find_str(rows_list, ["抖音人气榜"])),
                "star_rating": _find_float(rows_list, ["抖音星级"]),
                "impressions": _find_int(rows_list, ["页面曝光人数"]),
                "visits": _find_int(rows_list, ["页面访问人数"]),
                "purchases": _find_int(rows_list, ["购买人数"]),
                "verifications": _find_int(rows_list, ["核销人数"]),
                "bad_keywords": _extract_bad_keywords(opinion),
            }
        return {}

    def _has_data(metric_dict):
        """检查指标字典是否有非空数据（排除基础字段）"""
        skip_keys = {"store_id", "metric_date", "platform", "bad_keywords"}
        return any(v is not None and v != "" for k, v in metric_dict.items() if k not in skip_keys)

    # 上周指标
    for platform in ("meituan", "dianping", "douyin"):
        m = _make_platform_metrics(rows_prev, prev_end, platform)
        if _has_data(m):
            metrics.append(m)

    # 本周指标
    for platform in ("meituan", "dianping", "douyin"):
        m = _make_platform_metrics(rows_curr, curr_end, platform)
        if _has_data(m):
            metrics.append(m)

    # 分析意见
    if opinion:
        goals = _extract_goals(opinion)
        analysis = {
            "store_id": sid,
            "period_start": prev_start or prev_end,
            "period_end": curr_end,
            "analysis_opinion": opinion,
            "goals": goals,
        }

    return metrics, errors, analysis


# ==================== 主解析函数 ====================

def parse_excel(file_bytes: bytes, store_name_map: dict[str, UUID]) -> ParsedData:
    """
    解析周报 Excel 文件
    """
    result = ParsedData()

    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as e:
        result.errors.append(f"无法打开 Excel 文件: {str(e)}")
        return result

    sheet_names = wb.sheetnames
    store_sheet_names = ["大学城店", "涪陵店", "江津店", "微电园店", "奉节店", "綦江店"]

    # 1. 营业额
    if "营业额情况" in sheet_names:
        revs, errs, period = _parse_revenue_sheet(wb["营业额情况"], store_name_map)
        result.revenues.extend(revs)
        result.errors.extend(errs)
    else:
        result.errors.append("未找到 '营业额情况' sheet")
        period = (None, None)

    # 2. 套餐
    if "套餐数据" in sheet_names:
        pkgs, errs = _parse_package_sheet(wb["套餐数据"], store_name_map)
        result.packages.extend(pkgs)
        result.errors.extend(errs)
    else:
        result.errors.append("未找到 '套餐数据' sheet")

    # 3. 各门店指标和分析
    for sn in store_sheet_names:
        if sn in sheet_names:
            ms, errs, analysis = _parse_store_sheet(wb[sn], store_name_map, sn, period)
            result.metrics.extend(ms)
            result.errors.extend(errs)
            if analysis:
                result.analyses.append(analysis)

    wb.close()
    return result
